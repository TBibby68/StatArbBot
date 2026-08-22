import pandas as pd
from sqlalchemy import create_engine, text
from StatArbBot.config import engine_string
from openpyxl import load_workbook

engine = create_engine(engine_string)

print(f"Connected to: {engine.url}")

def on_minute_bar(timestamp, prices):
    # prices = {"JPM": 302.15, "BAC": 51.42}

    # 1. update rolling history
    # 2. calculate spread
    # 3. calculate z-score
    # 4. decide OPEN / CLOSE / NOTHING
    # 5. return a signal object
    pass

with engine.connect() as conn:

    # Check table exists
    exists = conn.execute(text("""
        SELECT EXISTS (
            SELECT 1
            FROM information_schema.tables
            WHERE table_name = 'backtesting_data_prices'
        );
    """)).scalar()

    print(f"Table exists: {exists}")

    if exists:

        # Read the whole table
        trades_df = pd.read_sql(
            "SELECT * FROM backtesting_data_prices",
            con=engine
        )

        datetime_columns = [
            col for col in trades_df.columns
            if col == "timestamp" or col.endswith("_last_update")
        ]

        for col in datetime_columns:
            trades_df[col] = (
                pd.to_datetime(trades_df[col], utc=True)
                .dt.tz_localize(None)
            )
        #trades_df["timestamp"] = trades_df["timestamp"].dt.tz_localize(None)

        output_file = r"C:\Users\tbibb\Downloads\backtesting_data_prices.xlsx"

        # Get the name of the first worksheet
        workbook = load_workbook(output_file)
        first_sheet_name = workbook.sheetnames[0]
        workbook.close()

        # Replace ONLY the first worksheet
        with pd.ExcelWriter(
            output_file,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            trades_df.to_excel(
                writer,
                sheet_name=first_sheet_name,
                index=False
            )

        print(
            f"Exported {len(trades_df)} rows to "
            f"'{first_sheet_name}' in '{output_file}'"
        )

    else:
        print("Table 'backtesting_data_prices' not found.")